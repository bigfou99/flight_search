import openpyxl
from openpyxl.utils import get_column_letter

import pandas as pd

def format_date(date):
    year = date.year
    month = str(date.month).zfill(2)  # Pad with zeros if necessary
    day = str(date.day).zfill(2)  # Pad with zeros if necessary
    return f"{year}-{month}-{day}"

def save_excel(obj, name):
    # Save to Excel
    df = pd.DataFrame(obj)
    df.to_excel(name, index=False)
    wb = openpyxl.load_workbook(name)

    # Open the default sheet
    sheet = wb['Sheet1']

    # Loop through the columns and set column width
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the changes
    wb.save(name)