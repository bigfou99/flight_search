import json
import requests
import pandas as pd
from datetime import datetime, timedelta
import brotli
import time
import random
import openpyxl
from openpyxl.utils import get_column_letter

simplified_itineraries = []



def format_date(date):
    year = date.year
    month = str(date.month).zfill(2)  # Pad with zeros if necessary
    day = str(date.day).zfill(2)  # Pad with zeros if necessary
    return f"{year}-{month}-{day}"

simplified_itineraries = []
error = False

ids = ["Station:airport:ORY", "Station:airport:CDG", "City:beirut_lb"]

# Simulate form data
form_data = {
    'nights': 7,
    'startDate': '2024-02-01',
    'endDate': '2024-02-29',
    "source": {
        "ids": [
            "City:beirut_lb"
        ]
    },
    "destination": {
        "ids": [
            "Station:airport:CDG",
            "Station:airport:ORY"
        ]
    }
}

nights = int(form_data['nights'])
start_date = datetime.strptime(form_data['startDate'], '%Y-%m-%d')
end_date = datetime.strptime(form_data['endDate'], '%Y-%m-%d')
loop_end_date = end_date - timedelta(days=nights)

d = start_date
while d <= loop_end_date:
    print(f"Scanning departure date: {d}")
    return_date = d + timedelta(days=nights)
    
    if return_date > end_date:
        break
    
    request_data = {
                "query": "query SearchReturnItinerariesQuery(\n  $search: SearchReturnInput\n  $filter: ItinerariesFilterInput\n  $options: ItinerariesOptionsInput\n) {\n  returnItineraries(search: $search, filter: $filter, options: $options) {\n    __typename\n    ... on AppError {\n      error: message\n    }\n    ... on Itineraries {\n      server {\n        requestId\n        environment\n        packageVersion\n        serverToken\n      }\n      metadata {\n        eligibilityInformation {\n          baggageEligibilityInformation {\n            topFiveResultsBaggageEligibleForPrompt\n            topFifteenResultsBaggageEligibleForPrompt\n            searchIsLongTrip\n            searchIsFamilyTrip\n            numberOfBags\n          }\n        }\n        ...AirlinesFilter_data\n        ...CountriesFilter_data\n        ...WeekDaysFilter_data\n        ...TravelTip_data\n        ...Sorting_data\n        ...MobileSorting_data\n        ...MobileSortingLink_data\n        ...PriceAlert_data\n        itinerariesCount\n        hasMorePending\n        missingProviders {\n          code\n        }\n        searchFingerprint\n        statusPerProvider {\n          provider {\n            id\n          }\n          errorHappened\n          errorMessage\n        }\n        topFiveOriginalItinerariesContainTHC\n        hasTier1MarketItineraries\n        sharedItinerary {\n          __typename\n          ... on ItineraryReturn {\n            ... on Itinerary {\n              __isItinerary: __typename\n              __typename\n              id\n              shareId\n              price {\n                amount\n                priceBeforeDiscount\n              }\n              priceEur {\n                amount\n              }\n              provider {\n                name\n                code\n                hasHighProbabilityOfPriceChange\n                contentProvider {\n                  code\n                }\n                id\n              }\n              bagsInfo {\n                includedCheckedBags\n                includedHandBags\n                hasNoBaggageSupported\n                hasNoCheckedBaggage\n                checkedBagTiers {\n                  tierPrice {\n                    amount\n                  }\n                  bags {\n                    weight {\n                      value\n                    }\n                  }\n                }\n                handBagTiers {\n                  tierPrice {\n                    amount\n                  }\n                  bags {\n                    weight {\n                      value\n                    }\n                  }\n                }\n                includedPersonalItem\n                personalItemTiers {\n                  tierPrice {\n                    amount\n                  }\n                  bags {\n                    weight {\n                      value\n                    }\n                    height {\n                      value\n                    }\n                    width {\n                      value\n                    }\n                    length {\n                      value\n                    }\n                  }\n                }\n              }\n              bookingOptions {\n                edges {\n                  node {\n                    token\n                    bookingUrl\n                    trackingPixel\n                    itineraryProvider {\n                      code\n                      name\n                      subprovider\n                      hasHighProbabilityOfPriceChange\n                      contentProvider {\n                        code\n                      }\n                      id\n                    }\n                    price {\n                      amount\n                    }\n                  }\n                }\n              }\n              travelHack {\n                isTrueHiddenCity\n                isVirtualInterlining\n                isThrowawayTicket\n              }\n            }\n            legacyId\n            outbound {\n              id\n              sectorSegments {\n                guarantee\n                segment {\n                  id\n                  source {\n                    localTime\n                    utcTime\n                    station {\n                      id\n                      legacyId\n                      name\n                      code\n                      type\n                      gps {\n                        lat\n                        lng\n                      }\n                      city {\n                        legacyId\n                        name\n                        id\n                      }\n                      country {\n                        code\n                        id\n                      }\n                    }\n                  }\n                  destination {\n                    localTime\n                    utcTime\n                    station {\n                      id\n                      legacyId\n                      name\n                      code\n                      type\n                      gps {\n                        lat\n                        lng\n                      }\n                      city {\n                        legacyId\n                        name\n                        id\n                      }\n                      country {\n                        code\n                        id\n                      }\n                    }\n                  }\n                  duration\n                  type\n                  code\n                  carrier {\n                    id\n                    name\n                    code\n                  }\n                  operatingCarrier {\n                    id\n                    name\n                    code\n                  }\n                  cabinClass\n                  hiddenDestination {\n                    city {\n                      name\n                      id\n                    }\n                    id\n                  }\n                  throwawayDestination {\n                    id\n                  }\n                }\n                layover {\n                  duration\n                  isBaggageRecheck\n                  isWalkingDistance\n                  transferDuration\n                  id\n                }\n              }\n              duration\n            }\n            inbound {\n              id\n              sectorSegments {\n                guarantee\n                segment {\n                  id\n                  source {\n                    localTime\n                    utcTime\n                    station {\n                      id\n                      legacyId\n                      name\n                      code\n                      type\n                      gps {\n                        lat\n                        lng\n                      }\n                      city {\n                        legacyId\n                        name\n                        id\n                      }\n                      country {\n                        code\n                        id\n                      }\n                    }\n                  }\n                  destination {\n                    localTime\n                    utcTime\n                    station {\n                      id\n                      legacyId\n                      name\n                      code\n                      type\n                      gps {\n                        lat\n                        lng\n                      }\n                      city {\n                        legacyId\n                        name\n                        id\n                      }\n                      country {\n                        code\n                        id\n                      }\n                    }\n                  }\n                  duration\n                  type\n                  code\n                  carrier {\n                    id\n                    name\n                    code\n                  }\n                  operatingCarrier {\n                    id\n                    name\n                    code\n                  }\n                  cabinClass\n                  hiddenDestination {\n                    city {\n                      name\n                      id\n                    }\n                    id\n                  }\n                  throwawayDestination {\n                    id\n                  }\n                }\n                layover {\n                  duration\n                  isBaggageRecheck\n                  isWalkingDistance\n                  transferDuration\n                  id\n                }\n              }\n              duration\n            }\n            stopover {\n              nightsCount\n              arrival {\n                type\n                city {\n                  name\n                  id\n                }\n                id\n              }\n              departure {\n                type\n                id\n              }\n              duration\n            }\n            lastAvailable {\n              seatsLeft\n            }\n            extendedFareOptionsPricing {\n              standardFarePrice {\n                amount\n              }\n              flexiFarePrice {\n                amount\n              }\n            }\n          }\n          id\n        }\n        kayakEligibilityTest {\n          containsKayakWithNewRules\n          containsKayakWithCurrentRules\n        }\n      }\n      itineraries {\n        __typename\n        ... on ItineraryReturn {\n          ... on Itinerary {\n            __isItinerary: __typename\n            __typename\n            id\n            shareId\n            price {\n              amount\n              priceBeforeDiscount\n            }\n            priceEur {\n              amount\n            }\n            provider {\n              name\n              code\n              hasHighProbabilityOfPriceChange\n              contentProvider {\n                code\n              }\n              id\n            }\n            bagsInfo {\n              includedCheckedBags\n              includedHandBags\n              hasNoBaggageSupported\n              hasNoCheckedBaggage\n              checkedBagTiers {\n                tierPrice {\n                  amount\n                }\n                bags {\n                  weight {\n                    value\n                  }\n                }\n              }\n              handBagTiers {\n                tierPrice {\n                  amount\n                }\n                bags {\n                  weight {\n                    value\n                  }\n                }\n              }\n              includedPersonalItem\n              personalItemTiers {\n                tierPrice {\n                  amount\n                }\n                bags {\n                  weight {\n                    value\n                  }\n                  height {\n                    value\n                  }\n                  width {\n                    value\n                  }\n                  length {\n                    value\n                  }\n                }\n              }\n            }\n            bookingOptions {\n              edges {\n                node {\n                  token\n                  bookingUrl\n                  trackingPixel\n                  itineraryProvider {\n                    code\n                    name\n                    subprovider\n                    hasHighProbabilityOfPriceChange\n                    contentProvider {\n                      code\n                    }\n                    id\n                  }\n                  price {\n                    amount\n                  }\n                }\n              }\n            }\n            travelHack {\n              isTrueHiddenCity\n              isVirtualInterlining\n              isThrowawayTicket\n            }\n          }\n          legacyId\n          outbound {\n            id\n            sectorSegments {\n              guarantee\n              segment {\n                id\n                source {\n                  localTime\n                  utcTime\n                  station {\n                    id\n                    legacyId\n                    name\n                    code\n                    type\n                    gps {\n                      lat\n                      lng\n                    }\n                    city {\n                      legacyId\n                      name\n                      id\n                    }\n                    country {\n                      code\n                      id\n                    }\n                  }\n                }\n                destination {\n                  localTime\n                  utcTime\n                  station {\n                    id\n                    legacyId\n                    name\n                    code\n                    type\n                    gps {\n                      lat\n                      lng\n                    }\n                    city {\n                      legacyId\n                      name\n                      id\n                    }\n                    country {\n                      code\n                      id\n                    }\n                  }\n                }\n                duration\n                type\n                code\n                carrier {\n                  id\n                  name\n                  code\n                }\n                operatingCarrier {\n                  id\n                  name\n                  code\n                }\n                cabinClass\n                hiddenDestination {\n                  city {\n                    name\n                    id\n                  }\n                  id\n                }\n                throwawayDestination {\n                  id\n                }\n              }\n              layover {\n                duration\n                isBaggageRecheck\n                isWalkingDistance\n                transferDuration\n                id\n              }\n            }\n            duration\n          }\n          inbound {\n            id\n            sectorSegments {\n              guarantee\n              segment {\n                id\n                source {\n                  localTime\n                  utcTime\n                  station {\n                    id\n                    legacyId\n                    name\n                    code\n                    type\n                    gps {\n                      lat\n                      lng\n                    }\n                    city {\n                      legacyId\n                      name\n                      id\n                    }\n                    country {\n                      code\n                      id\n                    }\n                  }\n                }\n                destination {\n                  localTime\n                  utcTime\n                  station {\n                    id\n                    legacyId\n                    name\n                    code\n                    type\n                    gps {\n                      lat\n                      lng\n                    }\n                    city {\n                      legacyId\n                      name\n                      id\n                    }\n                    country {\n                      code\n                      id\n                    }\n                  }\n                }\n                duration\n                type\n                code\n                carrier {\n                  id\n                  name\n                  code\n                }\n                operatingCarrier {\n                  id\n                  name\n                  code\n                }\n                cabinClass\n                hiddenDestination {\n                  city {\n                    name\n                    id\n                  }\n                  id\n                }\n                throwawayDestination {\n                  id\n                }\n              }\n              layover {\n                duration\n                isBaggageRecheck\n                isWalkingDistance\n                transferDuration\n                id\n              }\n            }\n            duration\n          }\n          stopover {\n            nightsCount\n            arrival {\n              type\n              city {\n                name\n                id\n              }\n              id\n            }\n            departure {\n              type\n              id\n            }\n            duration\n          }\n          lastAvailable {\n            seatsLeft\n          }\n          extendedFareOptionsPricing {\n            standardFarePrice {\n              amount\n            }\n            flexiFarePrice {\n              amount\n            }\n          }\n        }\n        id\n      }\n    }\n  }\n}\n\nfragment AirlinesFilter_data on ItinerariesMetadata {\n  carriers {\n    id\n    code\n    name\n  }\n}\n\nfragment CountriesFilter_data on ItinerariesMetadata {\n  stopoverCountries {\n    code\n    name\n    id\n  }\n}\n\nfragment MobileSortingLink_data on ItinerariesMetadata {\n  topResults {\n    best {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    cheapest {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    fastest {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    sourceTakeoffAsc {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    destinationLandingAsc {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n  }\n}\n\nfragment MobileSorting_data on ItinerariesMetadata {\n  topResults {\n    best {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    cheapest {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    fastest {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    sourceTakeoffAsc {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    destinationLandingAsc {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n  }\n}\n\nfragment PriceAlert_data on ItinerariesMetadata {\n  priceAlertExists\n  existingPriceAlert {\n    id\n  }\n  searchFingerprint\n  hasMorePending\n  priceAlertsTopResults {\n    best {\n      price {\n        amount\n      }\n    }\n    cheapest {\n      price {\n        amount\n      }\n    }\n    fastest {\n      price {\n        amount\n      }\n    }\n    sourceTakeoffAsc {\n      price {\n        amount\n      }\n    }\n    destinationLandingAsc {\n      price {\n        amount\n      }\n    }\n  }\n}\n\nfragment Sorting_data on ItinerariesMetadata {\n  topResults {\n    best {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    cheapest {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    fastest {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    sourceTakeoffAsc {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n    destinationLandingAsc {\n      __typename\n      duration\n      price {\n        amount\n      }\n      id\n    }\n  }\n}\n\nfragment TravelTip_data on ItinerariesMetadata {\n  travelTips {\n    __typename\n    ... on TravelTipRadiusMoney {\n      radius\n      params {\n        name\n        value\n      }\n      savingMoney: saving {\n        amount\n        currency {\n          id\n          code\n          name\n        }\n        formattedValue\n      }\n      location {\n        __typename\n        id\n        legacyId\n        name\n        slug\n      }\n    }\n    ... on TravelTipRadiusTime {\n      radius\n      params {\n        name\n        value\n      }\n      saving\n      location {\n        __typename\n        id\n        legacyId\n        name\n        slug\n      }\n    }\n    ... on TravelTipRadiusSome {\n      radius\n      params {\n        name\n        value\n      }\n      location {\n        __typename\n        id\n        legacyId\n        name\n        slug\n      }\n    }\n    ... on TravelTipDateMoney {\n      dates {\n        start\n        end\n      }\n      params {\n        name\n        value\n      }\n      savingMoney: saving {\n        amount\n        currency {\n          id\n          code\n          name\n        }\n        formattedValue\n      }\n    }\n    ... on TravelTipDateTime {\n      dates {\n        start\n        end\n      }\n      params {\n        name\n        value\n      }\n      saving\n    }\n    ... on TravelTipDateSome {\n      dates {\n        start\n        end\n      }\n      params {\n        name\n        value\n      }\n    }\n    ... on TravelTipExtend {\n      destination {\n        __typename\n        id\n        name\n        slug\n      }\n      locations {\n        __typename\n        id\n        name\n        slug\n      }\n      price {\n        amount\n        currency {\n          id\n          code\n          name\n        }\n        formattedValue\n      }\n    }\n  }\n}\n\nfragment WeekDaysFilter_data on ItinerariesMetadata {\n  inboundDays\n  outboundDays\n}\n",
                "variables": {
                    "search": {
                        "itinerary": {
                            "source": form_data["source"],
                            "destination": form_data["destination"],
                            "outboundDepartureDate": {
                                "start": format_date(d) + "T00:00:00",
                                "end": format_date(d) + "T23:59:59"
                            },
                            "inboundDepartureDate": {
                                "start": format_date(return_date) + "T00:00:00",
                                "end": format_date(return_date) + "T23:59:59"
                            }
                        },
                        "passengers": {
                            "adults": 1,
                            "children": 0,
                            "infants": 0,
                            "adultsHoldBags": [
                                0
                            ],
                            "adultsHandBags": [
                                0
                            ],
                            "childrenHoldBags": [],
                            "childrenHandBags": []
                        },
                        "cabinClass": {
                            "cabinClass": "ECONOMY",
                            "applyMixedClasses": True
                        }
                    },
                    "filter": {
                        "allowReturnFromDifferentCity": True,
                        "allowChangeInboundDestination": True,
                        "allowChangeInboundSource": True,
                        "allowDifferentStationConnection": True,
                        "enableSelfTransfer": True,
                        "enableThrowAwayTicketing": True,
                        "enableTrueHiddenCity": True,
                        "maxStopsCount": 0,
                        "transportTypes": [
                            "FLIGHT"
                        ],
                        "contentProviders": [
                            "KIWI"
                        ],
                        "flightsApiLimit": 25,
                        "limit": 30
                    },
                    "options": {
                        "sortBy": "QUALITY",
                        "mergePriceDiffRule": "INCREASED",
                        "currency": "usd",
                        "apiUrl": None,
                        "locale": "en",
                        "partner": "skypicker",
                        "partnerMarket": "en",
                        "affilID": "acquisition000performance000sem000google",
                        "storeSearch": False,
                        "searchStrategy": "REDUCED",
                        "abTestInput": {
                            "kayakABCTest": "REMOVE_KAYAK_COMPLETELY",
                            "applyRecommendedDestinationsSorting": False
                        },
                        "serverToken": None
                    }
                }
            }
    headers = {
        'Accept':'*/*',
        'Accept-Encoding':'gzip, deflate, br',
        'Accept-Language':'en-US,en;q=0.9',
        'Content-Type':'application/json',
        'Kw-Skypicker-Visitor-Uniqid':'b2bc59fe-9ac4-431a-9331-42c5f20ffc49',
        'Kw-Umbrella-Token':'d6d4f47ccdad80c2faa8c8edca2b8196d6845c5303151ea493e5405cdb97b284',
        'Origin':'https://www.kiwi.com',
        'Referer':'https://www.kiwi.com/',
        'Sec-Ch-Ua':'"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
        'Sec-Ch-Ua-Mobile':'?0',
        'Sec-Ch-Ua-Platform':'"Windows"',
        'Sec-Fetch-Dest':'empty',
        'Sec-Fetch-Mode':'cors',
        'Sec-Fetch-Site':'cross-site',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
    }
    
    try:
        response = requests.post('https://api.skypicker.com/umbrella/v2/graphql?featureName=SearchReturnItinerariesQuery', headers=headers, json=request_data)
        
        string_response = response.content.decode('utf-8')

        # Parse string to Python dictionary
        json_response = json.loads(string_response)
        
        original_itineraries = json_response['data']['returnItineraries']['itineraries']
        
        
        for itinerary in original_itineraries:
            simplified = {
                'Price Amount': itinerary['price']['amount'],
                'included Checked Bags': itinerary['bagsInfo']['includedCheckedBags'],
                'Included Hand Bags': itinerary['bagsInfo']['includedHandBags'],
                'Departure Time Out': itinerary['outbound']['sectorSegments'][0]['segment']['source']['localTime'],
                'Departure Time Return': itinerary['inbound']['sectorSegments'][0]['segment']['source']['localTime'],
                'Seats Left': itinerary['lastAvailable']['seatsLeft'],
                'Carrier Out': itinerary['outbound']['sectorSegments'][0]['segment']['carrier']['name'],
                'Carrier Return': itinerary['inbound']['sectorSegments'][0]['segment']['carrier']['name'],
                'Out Duration': str(round((itinerary['outbound']['duration']/3600), 2)) + " hours",
                'Return  Duration': str(round((itinerary['inbound']['duration']/3600), 2)) + " hours",
                'Out Arrival Time': itinerary['outbound']['sectorSegments'][0]['segment']['destination']['localTime'],
                'Return Arrival Time': itinerary['inbound']['sectorSegments'][0]['segment']['destination']['localTime'],
                'Standard Fare Price': itinerary['extendedFareOptionsPricing']['standardFarePrice']['amount'],
                'Flexi Fare Price': itinerary['extendedFareOptionsPricing']['flexiFarePrice']['amount']
            }
            
            simplified_itineraries.append(simplified)
            
    except Exception as e:
        error = True
        print(f"Error: {e}")
    
    d += timedelta(days=1)
    
    # Wait for x seconds before sending the next request (optional)
    random_float = random.uniform(1, 3)
    time.sleep(random_float)

if not error:
    # Sort by 'priceAmount' in ascending order
    sorted_itineraries = sorted(simplified_itineraries, key=lambda x: float(x['Price Amount']))
    
    # Save to Excel
    df = pd.DataFrame(sorted_itineraries)
    df.to_excel('Itineraries.xlsx', index=False)
    wb = openpyxl.load_workbook('Itineraries.xlsx')

    # Open the default sheet
    sheet = wb['Sheet1']

    # Loop through the columns and set column width
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the changes
    wb.save('Itineraries.xlsx')

    print("Done.")
